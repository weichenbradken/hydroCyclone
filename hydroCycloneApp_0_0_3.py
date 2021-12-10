import streamlit as st
import pandas as pd
import numpy as np
from PIL import Image, ImageDraw, ImageFont
import math
import os
import io
import streamlit.components.v1 as components
from openpyxl import load_workbook
#import plotly.graph_objects as go
from fpdf import FPDF
import base64
# SETTING PAGE CONFIG TO WIDE MODE
#st.set_page_config(layout="wide")
#image = Image.open('bradken.png')
#st.image(image, width = 50)

def CLEAR_TMP():
    if os.path.exists("resources/PDFtmp1.png"):
        os.remove("resources/PDFtmp1.png")
    elif os.path.exists("resources/PDFtmp2.png"):
        os.remove("resources/PDFtmp2.png")
    elif os.path.exists("fenji.out"):
        os.remove("fenji.out")
    return

def MASSBALANCE1(dryton, ovfSF, rcyload, oreDen, ovfWt, feedWt, ng_200, sandConc, ng_200_a, df1):
    # b2
    C2 = float(dryton)
    C5 = float(ovfWt)
    C8 = float(ovfSF)
    D5 = float(sandConc)
    D8 = float(ng_200_a)
    # calcu overflow concentration
    C4 = (100-C5)/C5
    C3 = C2*C4
    C6 = (C5/100)/(float(oreDen)+C5/100*(1-float(oreDen)))*100
    C7 = float(oreDen)/(C5/100+float(oreDen)*(1-C5/100))
    C9 = C2/C5*100
    C10 = C9/C7
    C11 = C10/3.6
    # calcu feed ore section
    D2 = C2*float(rcyload)/100
    D4 = (100-D5)/D5
    D3 = D2*D4
    D6 = (D5/100)/(float(oreDen)+D5/100*(1-float(oreDen)))*100
    D7 = float(oreDen)/(D5/100+float(oreDen)*(1-D5/100))
    D9 = D2/D5*100
    D10 = D9/D7
    D11 = D10/3.6
    # calcu sand concentration
    B2 = C2 + D2
    B3 = C3 + D3
    B8 = ((C8/100) + float(rcyload)/100*(D8/100))/(1+float(rcyload)/100)*100
    B9 = C9 + D9
    B5 = B2/B9*100
    B4 = (100-B5)/B5
    B6 = (B5/100)/(float(oreDen)+B5/100*(1-float(oreDen)))*100
    B7 = float(oreDen)/(B5/100+float(oreDen)*(1-B5/100))
    B10 = B9/B7
    B11 = B10/3.6


    # write data into csv
    df1['给矿'] = [B2, B3, B4, B5, B6, B7, B8, B9, B10, B11]
    df1['溢流'] = [C2, C3, C4, C5, C6, C7, C8, C9, C10, C11]
    df1['沉砂'] = [D2, D3, D4, D5, D6, D7, D8, D9, D10, D11]
    return df1

def MASSBALANCE2(dryton, ovfSF, rcyload, oreDen, ovfWt, feedWt, ng_200, sandConc, ng_200_a, df1):
    C2 = float(dryton)
    C5 = float(ovfWt)
    C8 = float(ovfSF)
    B5 = float(feedWt)
    D8 = float(ng_200_a)
    # calcu overflow concentration
    C4 = (100-C5)/C5
    C3 = C2*C4
    C6 = (C5/100)/(float(oreDen)+C5/100*(1-float(oreDen)))*100
    C7 = float(oreDen)/(C5/100+float(oreDen)*(1-C5/100))
    C9 = C2/C5*100
    C10 = C9/C7
    C11 = C10/3.6
    # calcu feed ore section
    D2 = C2*float(rcyload)/100
    B2 = C2 + D2
    B4 = (100-B5)/B5
    B3 = B2*B4
    B6 = (B5/100)/(float(oreDen)+B5/100*(1-float(oreDen)))*100
    B7 = float(oreDen)/(B5/100+float(oreDen)*(1-B5/100))
    B8 = ((C8/100) + float(rcyload)/100*(D8/100))/(1+float(rcyload)/100)*100
    B9 = B2 + B3
    B10 = B9/B7
    B11 = B10/3.6
    # calcu sand concentration
    D9 = B9 - C9
    D5 = D2/D9*100
    D6 = (D5/100)/(float(oreDen)+D5/100*(1-float(oreDen)))*100
    D7 = float(oreDen)/(D5/100+float(oreDen)*(1-D5/100))
    D10 = D9/D7
    D11 = D10/3.6
    D4 = (100-D5)/D5
    D3 = D2*D4
    # write data into csv
    df1['给矿'] = [B2, B3, B4, B5, B6, B7, B8, B9, B10, B11]
    df1['溢流'] = [C2, C3, C4, C5, C6, C7, C8, C9, C10, C11]
    df1['沉砂'] = [D2, D3, D4, D5, D6, D7, D8, D9, D10, D11]
    return df1

def CALC_P(df3, WK_NM, DIA, FEED_DIA, coeff_overFlwSld):
    # 计算工作压力
    singleCAP = df3['给矿'][8]/float(WK_NM)
    p00 = 3.141592653*pow(float(FEED_DIA)/2000,2)
    p0 = singleCAP/p00/3600
    p1 = pow(p0, 2)
    p2 = pow(float(FEED_DIA)/float(DIA),2)
    p3 = (pow(1.5*float(DIA)/(float(DIA)*float(coeff_overFlwSld)), 1.28)-1)/1000
    res = 10.7*df3['给矿'][5]*p1*p2*p3
    return res

def create_download_link(val, filename):
    b64 = base64.b64encode(val)  # val looks like b'...'
    return f'<a href="data:application/octet-stream;base64,{b64.decode()}" download="{filename}.xlsx">下载选型结果</a>'

def XLSX_WRITE():
    massBalancePara = pd.read_csv('mb_tpm1.csv', index_col=0)
    strucParameters = pd.read_csv('strucParaTMP.csv', index_col=0)
    ManufacturParam = pd.read_csv('manufCapTMP.csv', index_col=0)
    inputParam = pd.read_csv('inputPara.csv', index_col=0)
    sizeCheckPara = pd.read_csv('sizeCheckDF.csv', index_col=0)

    # read project info
    with open('projectInfo.out', 'r') as f:
        lines = f.readlines()
        if lines[0] is not None:
            mineName = lines[0].strip('\n')
        else:
            mineName = ' '
        if lines[0] is not None:
            projectName = lines[1].strip('\n')
        else:
            projectName = ' '
    f.close()

    # load template excel file
    wb = load_workbook(filename = "旋流器选型结果.xlsx")
    #Get the current Active Sheet
    ws = wb.active
    ## w = 3870 文档名称
    title = mineName + "选矿厂旋流器组选型计算说明书"
    ws['A2'] = title
    # 项目名称 - project name
    project_name = projectName
    ws['D3'] = project_name
    # 设备名称 - Equipment name
    equip_name = "NP-X" + str(int(strucParameters['Diameter'][0])) + "-" + str(int(ManufacturParam['setUnitNum'][0]) + int(strucParameters['BackUpNum'][0]))
    ws['D4'] = equip_name
    # 给矿量
    newFeed = round(inputParam['dryton'][0], 1)
    ws['D6'] = newFeed
    # 矿石密度
    oreDensity = inputParam['OreDensity'][0]
    ws['D7'] = oreDensity
    # 给矿浓度
    feedConcentration = round(massBalancePara['给矿'][3],1)
    ws['D8'] = feedConcentration
    # 给料粒度-200
    feedSize_200 = round(massBalancePara['给矿'][6],1)
    ws['H7'] = feedSize_200
    # 溢流细度-200
    overFlow_200 = inputParam['ovfSF'][0]
    ws['H8'] = overFlow_200
    # 返矿比
    recycelLoad = inputParam['rcyload'][0]
    ws['H6'] = recycelLoad
    # 物料平衡 - 进料
    feed_r1 = round(massBalancePara['给矿'][0],1)
    feed_r2 = round(massBalancePara['给矿'][1],1)
    feed_r3 = round(massBalancePara['给矿'][3],1)
    feed_r4 = round(massBalancePara['给矿'][5],2)
    feed_r5 = round(massBalancePara['给矿'][8],1)
    ws['F11'] = feed_r1
    ws['F12'] = feed_r2
    ws['F13'] = feed_r3
    ws['F14'] = feed_r4
    ws['F15'] = feed_r5
    # 物料平衡 - 溢流
    oflw_r1 = round(massBalancePara['溢流'][0],1)
    oflw_r2 = round(massBalancePara['溢流'][1],1)
    oflw_r3 = round(massBalancePara['溢流'][3],1)
    oflw_r4 = round(massBalancePara['溢流'][5],2)
    oflw_r5 = round(massBalancePara['溢流'][8],1)
    ws['G11'] = oflw_r1
    ws['G12'] = oflw_r2
    ws['G13'] = oflw_r3
    ws['G14'] = oflw_r4
    ws['G15'] = oflw_r5
    # 物料平衡 - 底流
    sand_r1 = round(massBalancePara['沉砂'][0],1)
    sand_r2 = round(massBalancePara['沉砂'][1],1)
    sand_r3 = round(massBalancePara['沉砂'][3],1)
    sand_r4 = round(massBalancePara['沉砂'][5],2)
    sand_r5 = round(massBalancePara['沉砂'][8],1)
    ws['H11'] = sand_r1
    ws['H12'] = sand_r2
    ws['H13'] = sand_r3
    ws['H14'] = sand_r4
    ws['H15'] = sand_r5
    # 旋流器直径
    cyclone_dia = int(strucParameters['Diameter'][0])
    ws['H17'] = cyclone_dia
    # 校核单台生产能力
    cyclone_D = strucParameters['Diameter'][0]/10
    cyclone_di = strucParameters['FeedDiameter'][0]/10
    cyclone_do = round(strucParameters['OverFlowWidth'][0]/10,1)
    cyclone_P = strucParameters['SetPressure'][0]
    cyclone_qm = round(ManufacturParam['UnitCap'][0],1)
    ws['H22'] = cyclone_D
    ws['H23'] = cyclone_di
    ws['H24'] = cyclone_do
    ws['H25'] = cyclone_P
    ws['H26'] = cyclone_qm
    # 旋流器数量
    cyclone_num = str(int(ManufacturParam['setUnitNum'][0]))
    ws['H29'] = cyclone_num
    # 计算和校核分级粒度
    cyclone_check_D = strucParameters['Diameter'][0]/10
    cyclone_check_do = strucParameters['OverFlowWidth'][0]/10
    cyclone_check_di = strucParameters['FeedDiameter'][0]/10
    cyclone_check_pm = round(massBalancePara['给矿'][5],2)
    cyclone_check_um = round(sizeCheckPara['viscosity'][0],5)
    cyclone_check_delta = inputParam['OreDensity'][0]
    cyclone_check_Pm = strucParameters['SetPressure'][0]
    cyclone_check_theta = strucParameters['Angle'][0]
    cyclone_check_dm = round(sizeCheckPara['fenJiLiDu'][0],1)
    ws['H34'] = cyclone_check_D
    ws['H35'] = cyclone_check_do
    ws['H36'] = cyclone_check_di
    ws['H37'] = cyclone_check_pm
    ws['H38'] = cyclone_check_um
    ws['H39'] = cyclone_check_delta
    ws['H40'] = cyclone_check_Pm
    ws['H41'] = cyclone_check_theta
    ws['H42'] = cyclone_check_dm
    # 结论
    summry = "结论：分级粒度" + str(round(sizeCheckPara['fenJiLiDu'][0],1)) + "<" + str(inputParam['fenji_200'][0]) + "μm (-200目" + str(round(massBalancePara['溢流'][6],1)) + "%),满足设计要求"
    ws['A43'] = summry
    # 与设备名称一致
    ws['G45'] = equip_name
    # 旋流器组型号
    ops_cyclone_qty = int(ManufacturParam['setUnitNum'][0])
    spr_cyclone_qty = int(strucParameters['BackUpNum'][0])
    ws['H47'] = ops_cyclone_qty
    ws['H48'] = spr_cyclone_qty
    # 旋流器结构参数 
    cyclone_struc_p1 = str(int(strucParameters['FeedDiameter'][0]))
    cyclone_struc_p2 = str(strucParameters['OverFlowWidth'][0])
    cyclone_struc_p3 = str(int(strucParameters['SandDia'][0]))
    cyclone_struc_p4 = str(strucParameters['SetPressure'][0])
    ws['H51'] = cyclone_struc_p1
    ws['H52'] = cyclone_struc_p2
    ws['H54'] = cyclone_struc_p3
    ws['H56'] = cyclone_struc_p4
    wb.save("选型结果.xlsx")
    return


def main():
    st.set_page_config(page_title='水力旋流器选型', initial_sidebar_state = 'auto')
    #page_icon = favicon,
    st.markdown(
            f"""
            <style>
                .reportview-container .main .block-container{{
                    max-width: 1350px;
                    padding-top: 1rem;
                    padding-right: 1rem;
                    padding-left: 1rem;
                    padding-bottom: 1rem;
                }}

            </style>
            """,
            unsafe_allow_html=True,
        )

    #
    #            .reportview-container .main {{
    #                color: {COLOR};
    #                background-color: {BACKGROUND_COLOR};
    #            }}

    # Clear all tmp files
    CLEAR_TMP()
    # main section
    #head_c1, head_c2, head_c3, head_c4 = st.columns(4)
    #head_c3.subheader('水力旋流器选型软件')
    st.markdown("<h2 style='text-align: center; color: black;'>水力旋流器选型软件</h2>", unsafe_allow_html=True)
    st.sidebar.subheader("参考图表及曲线")
    image1 = Image.open('fig1.jpg')
    sizeCheck1 = st.sidebar.expander("沉砂能力-沉砂口直径校核", True)
    sizeCheck1.image(image1)
    image2 = Image.open('fig2.jpg')
    sizeCheck2 = st.sidebar.expander("溢流细度-沉砂细度校核", True)
    sizeCheck2.image(image2)
    image3 = Image.open('structparameters.png')
    sizeCheck3 = st.sidebar.expander("结构参数参照表", True)
    sizeCheck3.image(image3)
    image4 = Image.open('sizeCheck.png')
    sizeCheck4 = st.sidebar.expander("分级粒度校核表", True)
    sizeCheck4.image(image4)

    

    original_title_1 = '<p style="font-family:Times New Roman; color:Black; font-size: 16px; font-weight: bold;">项目信息</p>'
    st.markdown(original_title_1, unsafe_allow_html=True)
    projectInfo = st.expander(" ", True)
    with projectInfo:
        #original_title_1 = '<p style="font-family:Times New Roman; color:Black; font-size: 16px; font-weight: bold;">项目信息</p>'
        #projectInfo.markdown(original_title_1, unsafe_allow_html=True)
        r1col1, r1col2, r1col3 = projectInfo.columns(3)
        tmpPI1 = r1col1.text_input('客户名称')
        tmpPI2 = r1col2.text_input('项目名称')
        r1col3.text_input('联 系 人')
        r2col1, r2col2, r2col3 = projectInfo.columns(3)
        r2col1.text_input('Email')
        r2col2.text_input('电 话')
        r2col3.text_input('微 信')
        with open('projectInfo.out', 'w') as f:
            f.write(tmpPI1)
            f.write('\n')
            f.write(tmpPI2)
            f.write('\n')
        f.close()



    original_title_2 = '<p style="font-family:Times New Roman; color:Black; font-size: 16px; font-weight: bold;">旋流器用途</p>'
    st.markdown(original_title_2, unsafe_allow_html=True)
    application = st.expander("", True)
    with application:
        application.radio('请选择旋流器应用场景', ["分级", "脱水", "胶泥", "除杂", "浓缩", "筑坝"])
        application.text_input('其他-请说明')
        application.write('<style>div.row-widget.stRadio > div{flex-direction:row;}</style>', unsafe_allow_html=True)



    original_title_3 = '<p style="font-family:Times New Roman; color:Black; font-size: 16px; font-weight: bold;">工况及要求指标</p>'
    st.markdown(original_title_3, unsafe_allow_html=True)
    parameters = st.expander("", True)
    with parameters:
        #parameters.container("1.旋流器给矿条件")
        parameters.write("1. 旋流器给矿条件")
        # first row
        r1col1, r1col2, r1col3 = parameters.columns(3)
        r1col1.text_input('处理矿石名称')
        oreDen = r1col2.text_input('矿 石 密 度 - [t/m³]', '0')
        feedWt = r1col3.text_input('给矿质量浓度 - [%]', '0')
        # second row
        r2col1, r2col2, r2col3 = parameters.columns(3)
        r2col1.text_input('矿浆温度 - [ᵒC]')
        ng_200 = r2col1.text_input('给矿-200目含量 - [%]', '0')
        dryton = r2col2.text_input('干矿处理量 - [t/h]', '0')
        rcyload = r2col3.text_input('循环负荷 - [%]', '0')
        # third row
        
        # seconds container
        parameters.write("2. 要求旋流器的工作效果")
        # first row
        rr1col1, rr1col2, rr1col3 = parameters.columns(3)
        ovfSF = rr1col1.text_input('溢流细度(-200目含量) - [%]', '0')
        fenji_200 = rr1col2.text_input("分级细度 - [um]", '0')
        ovfWt = rr1col3.text_input('溢流质量浓度 - [%]', '0')
        ng_200_a = rr1col1.text_input("沉砂细度-200目含量 - [%]", '0')
        sandConc = rr1col2.text_input("沉砂质量浓度 - [%]", '0')
        
        with open('fenji.out', 'w') as f:
            f.write(fenji_200)
        f.close()
        inputPara = pd.DataFrame({'OreDensity':float(oreDen), 'feedWt':float(feedWt), 'ng_200':float(ng_200), \
                                  'dryton':float(dryton), 'rcyload':float(rcyload), 'ovfSF':float(ovfSF), \
                                  'ovfWt':float(ovfWt), 'ng_200_a':float(ng_200_a), 'sandConc':float(sandConc), \
                                  'fenji_200':float(fenji_200)}, index=list('A'))
        inputPara.to_csv('inputPara.csv')


    # 物料平衡计算
    # 分为两列
    original_title_4 = '<p style="font-family:Times New Roman; color:Black; font-size: 16px; font-weight: bold;">物料平衡计算</p>'
    st.markdown(original_title_4, unsafe_allow_html=True)
    massbalance = st.expander('', True)
    with massbalance:
        #massbalance.write("工作条件")
        # 显示已输入参数
        #MB_r1col1, MB_r1col2, MB_r1col3 = massbalance.columns(3)
        #MB_r1col1.metric(label="干矿处理量 - [t/h]", value=dryton)
        #MB_r1col1.metric(label="溢流细度(-200目含量) - [%]", value=ovfSF)
        #MB_r1col2.metric(label="循环负荷 - [%]", value=rcyload)
        #MB_r1col2.metric(label="矿石密度 - [t/m³]", value=oreDen)
        #MB_r1col3.metric(label="溢流质量浓度 - [%]", value=ovfWt)
        #MB_r1col3.metric(label="给矿重量浓度 - [%]", value=feedWt)
        
        # 显示物料平衡
        #massbalance.markdown("--------------------------------------------")
        # 下面显示pandas dataframe. 根据用户选择的不同
        #mssBcol1, mssBcol2 = massbalance.columns(2)
        #with mssBcol1.form(key = "my_form1"):
        #    mssBcol1.subheader("物料平衡计算一：")
        #    mssBcol1.write("已知    沉砂质量浓度-%  及  沉砂细度-200目含量-%")
        #    submitted1 = mssBcol1.button(label = '计算物料平衡1')
            
        #with mssBcol2.form(key = "my_form2"):
        #    #feedConc = mssBcol2.text_input("给矿质量浓度 - [%]")
        #    #ng_200_b = mssBcol2.text_input("细度-200目含量2 - [%]")
        #    mssBcol2.subheader("物料平衡计算二：")
        #    mssBcol2.write("已知    给矿质量浓度-%  及  沉砂细度-200目含量-%")
        #    submitted2 = mssBcol2.button(label = '计算物料平衡2')
        #mssBcol1, mssBcol2 = massbalance.columns(2)
        submitted = massbalance.radio("", ('已知沉砂质量浓度', '已知给矿质量浓度'))
        #mssBcol2.markdown('__________________________')
        mb_button = massbalance.button(label = '计算物料平衡')

        df1 = pd.read_csv('mb_tpm1.csv', index_col=0)
        if mb_button and submitted == '已知沉砂质量浓度':
            #if 
            df1 = MASSBALANCE1(dryton, ovfSF, rcyload, oreDen, ovfWt, feedWt, ng_200, sandConc, ng_200_a, df1)
            #st.table(df1.style.format("{:.2f}"))
            #df1.to_csv('mb_tpm1.csv')
        elif mb_button and submitted == '已知给矿质量浓度':
            #df1 = pd.read_csv('massBalance_1.csv', index_col=0)
            df1 = MASSBALANCE2(dryton, ovfSF, rcyload, oreDen, ovfWt, feedWt, ng_200, sandConc, ng_200_a, df1)
        st.table(df1.style.format("{:.2f}"))
        df1.to_csv('mb_tpm1.csv')
        


    # 结构参数
    original_title_5 = '<p style="font-family:Times New Roman; color:Black; font-size: 16px; font-weight: bold;">结构参数计算</p>'
    st.markdown(original_title_5, unsafe_allow_html=True)
    strucPara = st.expander("", True)
    with strucPara:

        # read previous cal results for this section
        df3 = pd.read_csv('mb_tpm1.csv', index_col=0)
        # 台数参数
        SP_r1col1, SP_r1col2, SP_r1col3, SP_r1col4 = strucPara.columns(4)
        WK_NM = SP_r1col1.text_input("工件台数", 1)
        BP_NM = SP_r1col2.text_input("备用台数", 0)
        DEG = SP_r1col3.text_input("锥    角 - [度]", 20)
        if  float(WK_NM) == 1:
            singleUnit = 1
            SP_r1col4.metric(label="单台需能力 - [m³/h]", value=singleUnit)
        elif float(WK_NM) > 1:
            singleUnit = df3['给矿'][8]/float(WK_NM)
            SP_r1col4.metric(label="单台需能力 - [m³/h]", value=round(singleUnit, 1))
        
        #print(singleUnit)
        # 根据计算压力为0.08 MPa计算直径。
        #strucPara.write("根据初始工作压力为：0.08 MPa 以及 以上单台所需处理能力计算所得直径为：")
        initialD = 1.95*pow(df3['给矿'][8]/float(WK_NM), 0.5)*pow(df3['给矿'][5], 0.25)*1.56691796
        strucPara.metric("推荐初始直径 - [cm]", round(initialD, 1))
        
        # 工作压力
        strucPara.markdown('-------------------------------------------------------------------')
        strucPara.markdown("设定结构参数：")
        SPPP_r1col1, SPPP_r1col2, SPPP_r1col3 = strucPara.columns(3)
        DIA = SPPP_r1col1.text_input("型号(直径) - [mm]", 1)
        #VALV_DIA = SPPP_r1col2.text_input("阀门直径 - [mm]", 1)
        FEED_DIA = SPPP_r1col3.text_input("给 矿 口 - [mm] (查表获得)", 1)

        overFlwSld_r1col1, overFlwSld_r1col2, overFlwSld_r1col3 = strucPara.columns(3)
        coeff_overFlwSld = overFlwSld_r1col1.slider('请选择溢流口与直径间的系数?', 0.36, 0.42, 0.38, 0.02)
        overFlwSld_r1col3.write("                        ")
        OFLW = int(float(coeff_overFlwSld)*float(DIA))
        overFlwSld_r1col3.metric("溢 流 口 - [mm]", OFLW)
    
        #-----------------------------------------------
        sanDia_r1col1, sanDia_r1col2, sanDia_r1col3 = strucPara.columns(3)
        coeff_sanDia = sanDia_r1col1.slider('请选择沉砂咀计算系数?', 4, 20, 12, 2)
        coeff_height = sanDia_r1col1.slider('请选择溢流管插入深度计算系数?', 0.4, 0.6, 0.45, 0.05)

        #if  df3['溢流'][6] > 75:
        CALC_SAND_DIA = pow(df3['沉砂'][8]/df3['溢流'][8]*pow(OFLW,3)/float(coeff_sanDia), 0.33333333)
        #else:
            #print(OFLW)
        #    calcSandTmp1 = pow(float(DIA)*0.42, 3)
        #    #print(calcSandTmp1)
        #    CALC_SAND_DIA = pow(df3['沉砂'][8]/df3['溢流'][8]*calcSandTmp1/float(coeff_sanDia), 0.33333333)
        sanDia_r1col3.write("                         ")
        sanDia_r1col3.metric("沉 砂 咀 - [mm]", int(CALC_SAND_DIA))
        sanDia_r1col3.write("                         ")
        OVFL_DEP = sanDia_r1col3.metric("溢流管插入深度 - [mm]", int(float(DIA)*float(coeff_height)))
        
        
        
        
        
        #  # 计算工作压力 与 人为设置工作压力
        SPP_r1col1, SPP_r1col2, SPP_r1col3 = strucPara.columns(3)
        if DIA is not None and FEED_DIA is not None:
            calcP = CALC_P(df3, WK_NM, DIA, FEED_DIA, coeff_overFlwSld)
            SPP_r1col1.metric(label="计算工作压力 - [MPa]", value=round(calcP, 3))
        else:
            SPP_r1col1.metric(label="计算工作压力 - [MPa]", value='待计算')
        setP = SPP_r1col3.text_input("设定工作压力 - [MPa]", 0.08)
        #SPP_r1col1.latex(r''' 
        #    \Delta P_{m} = 10.7 \rho_{m} V_{i}^{2} \left(\frac{d_{i}}{D}\right)^{2} \lbrack\left( 1.5\frac{D}{d_{o}} \right)^{1.28}-1\rbrack
        #    ''')
        #SPP_r1col2.metric(label="计算直径 - [cm]", value=81)
        #SPP_r1col2.latex(r''' 
        #    D = 1.95 q_{m}^{0.5} \rho_{m}^{0.25} \Delta P_{m}^{-0.25}
        #    ''')
        if setP == '1':
            SPP_r1col3.metric(label="校核直径 - [cm]", value='待计算')
        else:
            calcDia = 1.95*pow(df3['给矿'][8]/float(WK_NM), 0.5)*pow(df3['给矿'][5], 0.25)*pow(float(setP), -0.25)/1.2
            SPP_r1col3.metric(label="计算直径 - [cm]", value=round(calcDia, 1))

        #strucParaDF = pd.DataFrame([float(setP), float(DIA), float(FEED_DIA), float(OFLW)], index=list('1'), columns=list('ABCD'))
        strucParaDF = pd.DataFrame({'SetPressure':float(setP), 'Diameter':float(DIA), 'FeedDiameter':float(FEED_DIA), 'OverFlowWidth':OFLW, 'SandDia':CALC_SAND_DIA, 'SingleUnitCap':singleUnit, 'Angle':float(DEG), 'BackUpNum':float(BP_NM)}, index=list('A'))
        strucParaDF.to_csv('strucParaTMP.csv')

#   结构参数
    original_title_6 = '<p style="font-family:Times New Roman; color:Black; font-size: 16px; font-weight: bold;">生产能力计算</p>'
    st.markdown(original_title_6, unsafe_allow_html=True)
    manufCap = st.expander("", True)
    with manufCap:
        # load dfs
        csvFLAG = 0
        if os.path.isfile("mb_tpm1.csv") and os.path.isfile("strucParaTMP.csv"):
            manufCapTmpDf1 = pd.read_csv('mb_tpm1.csv', index_col=0)
            manufCapTmpDf2 = pd.read_csv('strucParaTMP.csv', index_col=0)
            csvFLAG = 1
        #print(DIA)
        #print(OFLW)
        #print(FEED_DIA)
        cycloneOPT = manufCap.radio("请选择旋流器类型", ('标准旋流器', '高效旋流器'))

        MC_col1, MC_col2, MC_col3 = manufCap.columns(3)
        if csvFLAG == 1 and cycloneOPT== '标准旋流器':
            tmp0 = pow(1.5*manufCapTmpDf2['Diameter'][0]/manufCapTmpDf2['OverFlowWidth'][0], 1.28)
            tmp1 = manufCapTmpDf2['SetPressure'][0]/(manufCapTmpDf1['给矿'][5]*(tmp0-1))
            tmp2 = 2.69*manufCapTmpDf2['Diameter'][0]/10*manufCapTmpDf2['FeedDiameter'][0]/10*pow(tmp1, 0.5)
            ManuCAP = tmp2
            MC_col1.metric(label="单台处理能力 - [m³/h]", value=round(ManuCAP, 1))

            unitNUM = manufCapTmpDf1['给矿'][8]/tmp2 #tmp2
            MC_col2.metric(label="数量", value=round(unitNUM, 1))
            #MC_col2.metric(label="单台处理能力 [高效旋流器] - [m³/h]", value=round(ManuCAP*1.25, 1))
        elif csvFLAG == 1 and cycloneOPT== '高效旋流器':
            tmp0 = pow(1.5*manufCapTmpDf2['Diameter'][0]/manufCapTmpDf2['OverFlowWidth'][0], 1.28)
            tmp1 = manufCapTmpDf2['SetPressure'][0]/(manufCapTmpDf1['给矿'][5]*(tmp0-1))
            tmp2 = 2.69*manufCapTmpDf2['Diameter'][0]/10*manufCapTmpDf2['FeedDiameter'][0]/10*pow(tmp1, 0.5)
            ManuCAP = tmp2*1.25
            #MC_col1.metric(label="单台处理能力 - [m³/h]", value=round(ManuCAP, 1))
            MC_col1.metric(label="单台处理能力 - [m³/h]", value=round(ManuCAP, 1))
            unitNUM = manufCapTmpDf1['给矿'][8]/ManuCAP #tmp2
            MC_col2.metric(label="数量", value=round(unitNUM, 1))
        else:
            ManuCAP = 1
            MC_col1.metric(label="单台处理能力 - [m³/h]", value=ManuCAP)
            #MC_col2.metric(label="单台处理能力 [高效旋流器] - [m³/h]", value=round(ManuCAP*1.25, 1))
        #else:
        #    ManuCAP = MC_col1.metric(label="计算生产能力 - [m³/h]", value=1)
        #    print(singleUnit)

        if ManuCAP > manufCapTmpDf2['SingleUnitCap'][0]:
            MC_col1.success("符合计算条件")
        else:
            MC_col1.error("计算生产能力 < 单台处理能力, 请检查设置!")
        #MC_col1.latex(r''' 
        #    q_{m} = 2.69 D d_{i} ( \, \frac{\Delta P_{m}}{\rho_{m} [ \, ( \, 1.5\frac{D}{d_{0}} ) \,^{1.28} - 1 ] \, } ) \,^{0.5}
        #    ''')
        #if csvFLAG == 1:

        #else:
        #    MC_col2.metric(label="数量", value='待计算')
        setUnitNum = MC_col3.text_input("设定工作台数", '0.1')
        # manufacture_cap
        manufCapDF = pd.DataFrame({'UnitCap':float(ManuCAP), 'setUnitNum':float(setUnitNum)}, index=list('A'))
        manufCapDF.to_csv('manufCapTMP.csv')

    #结构参数
    original_title_7 = '<p style="font-family:Times New Roman; color:Black; font-size: 16px; font-weight: bold;">分级粒度校核</p>'
    st.markdown(original_title_7, unsafe_allow_html=True)
    sizeClss = st.expander("", True)
    with sizeClss:
        # load csv
        sizeClssCsvFLAG = 0
        if os.path.isfile("mb_tpm1.csv") and os.path.isfile("strucParaTMP.csv") and os.path.isfile("manufCapTMP.csv"):
            sizeClssTmpDf1 = pd.read_csv('mb_tpm1.csv', index_col=0)
            sizeClssTmpDf2 = pd.read_csv('strucParaTMP.csv', index_col=0)
            sizeClssTmpDf3 = pd.read_csv('manufCapTMP.csv', index_col=0)
            sizeClssCsvFLAG = 1
        SC_col1, SC_col2, SC_col3 = sizeClss.columns(3)
        if sizeClssCsvFLAG == 1:
            szClsTmp1 = math.exp(16.6*sizeClssTmpDf1['给矿'][4]/100)
            szClsTmp2 = 10.05*pow(sizeClssTmpDf1['给矿'][4]/100,2)
            szClsTmp3 = 0.001*(1+2.5*sizeClssTmpDf1['给矿'][4]/100+szClsTmp2+0.00273*szClsTmp1)
            #SC_col1.metric(label="粘度 - [Pa.s]", value=round(szClsTmp3, 5))
        #else:
            #SC_col1.metric(label="粘度 - [Pa.s]", value='待计算')

        if sizeClssCsvFLAG == 1:
            szClsTmp1_00 = math.tan(sizeClssTmpDf2['Angle'][0]/2*3.141593/180)
            szClsTmp1_01 = (float(oreDen) - sizeClssTmpDf1['给矿'][5])*(3*sizeClssTmpDf2['Diameter'][0]/10-2*sizeClssTmpDf2['OverFlowWidth'][0]/10)*pow(sizeClssTmpDf2['SetPressure'][0], 0.5)
            szClsTmp1_02 = sizeClssTmpDf2['FeedDiameter'][0]/10*pow(sizeClssTmpDf1['给矿'][5], 0.5)
            szClsTmp1_03 = pow(sizeClssTmpDf2['OverFlowWidth'][0]/10, 0.64)
            szClsTmp1_04 = pow(sizeClssTmpDf2['Diameter'][0]/10, 0.36)
            szClsTmp1_05 = 1815*pow(szClsTmp1_04*szClsTmp1_03*szClsTmp1_02*szClsTmp3/szClsTmp1_01*szClsTmp1_00, 0.5)
            SC_col1.metric(label="分级粒度 - [um]", value=round(szClsTmp1_05, 1))

            # check error or sucess
            fenji_f = open("fenji.out", 'r')
            tmp_fenji = float(fenji_f.read())
            fenji_f.close()
            #print(tmp_fenji)
            if szClsTmp1_05 < tmp_fenji:
                SC_col1.success("计算分级细度度小于要求分级细度，满足设计要求")
            else:
                SC_col1.error("计算分级细度度大于要求分级细度，不满足设计要求")

        else:
            SC_col1.metric(label="分级粒度 - [um]", value='待计算')
        # save pdf
        sizeCheckDF = pd.DataFrame({'viscosity':szClsTmp3, 'fenJiLiDu':szClsTmp1_05}, index=list('A'))
        sizeCheckDF.to_csv('sizeCheckDF.csv')




    #   结构参数
    original_title_8 = '<p style="font-family:Times New Roman; color:Black; font-size: 16px; font-weight: bold;">沉砂能力校核</p>'
    st.markdown(original_title_8, unsafe_allow_html=True)
    sandsink = st.expander("", True)
    with sandsink:
        #
        SS_col1, SS_col2, SS_col3 = sandsink.columns(3)
        sandsinkFlag = 0
        if os.path.isfile("mb_tpm1.csv") and os.path.isfile("strucParaTMP.csv"):
             sandsinkDF1 = pd.read_csv('mb_tpm1.csv', index_col=0)
             sandsinkDF2 = pd.read_csv('strucParaTMP.csv', index_col=0)
             sandsinkFlag = 1
        if setUnitNum == '0.1' and sandsinkFlag==0 :
            SS_col1.metric(label="设计沉砂能力 - t/cm².h", value='待计算')
            SS_col3.metric(label="排口比", value='待计算')
            sandsink.metric(label="分级效率 - [%]", value='待计算')
        else:
            CONV_TMP = 1.27*sandsinkDF1['沉砂'][0]/float(setUnitNum)/pow(sandsinkDF2['SandDia'][0], 2)*100
            SS_col1.metric(label="设计沉砂能力 - [t/cm².h]", value=round(CONV_TMP, 1))
            if CONV_TMP < 0.5 or CONV_TMP > 2.5:
                SS_col1.error("计算错误, 请再次校验输入参数!")
            else:
                SS_col1.success("计算属于合理范围!")
            DiaRatioTMP = sandsinkDF2['SandDia'][0]/sandsinkDF2['OverFlowWidth'][0]
            SS_col3.metric(label="排口比", value=round(DiaRatioTMP, 2))
            if DiaRatioTMP < 0.33 or DiaRatioTMP > 0.58:
                SS_col3.error("计算错误, 请再次校验输入参数!")
            else:
                SS_col3.success("计算属于合理范围!")
            #sandsink.metric(label="分级效率 - [%]", value=round(sandsinkDF1['溢流'][6]*(sandsinkDF1['给矿'][6]-sandsinkDF1['沉砂'][6])/(sandsinkDF1['给矿'][6]*(sandsinkDF1['溢流'][6]-sandsinkDF1['沉砂'][6]))*100, 1))

    original_title_9 = '<p style="font-family:Times New Roman; color:Black; font-size: 16px; font-weight: bold;">分级效率计算</p>'
    st.markdown(original_title_9, unsafe_allow_html=True)
    fenjiEffcny = st.expander("", True)
    with fenjiEffcny:
        #
        fjSS_col1, fjSS_col2, fjSS_col3 = fenjiEffcny.columns(3)
        fjSSFlag = 0
        if os.path.isfile("mb_tpm1.csv") and os.path.isfile("strucParaTMP.csv"):
            fjSSFlagDF1 = pd.read_csv('mb_tpm1.csv', index_col=0)
            fjSSFlagDF2 = pd.read_csv('strucParaTMP.csv', index_col=0)
            fjSSFlag = 1
            fjSS_col1.metric(label="分级效率 - [%]", value=round(fjSSFlagDF1['溢流'][6]*(fjSSFlagDF1['给矿'][6]-fjSSFlagDF1['沉砂'][6])/(fjSSFlagDF1['给矿'][6]*(fjSSFlagDF1['溢流'][6]-fjSSFlagDF1['沉砂'][6]))*100, 1))
        else:
            fjSS_col1.metric(label="分级效率 - [%]", value="待计算")


    #   参考曲线及数据表
    exportPDF = st.button(label = '导出计算结果')
    if exportPDF:
        XLSX_WRITE()
        #IMAGE_WRITE()
        #pdf = FPDF()
        #imagelist = ["resources/PDFtmp1.png", "resources/PDFtmp2.png"]
        # imagelist is the list with all image filenames
        #for image in imagelist:
        #    pdf.add_page()
        #    pdf.image(image, 0, 0, 210, 297)
        wb = load_workbook(filename = "选型结果.xlsx")
        ws = wb.active
        towrite = io.BytesIO()
        wb.save(towrite)
        towrite.seek(0)
        b64 = base64.b64encode(towrite.read()).decode()
        html = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="选型结果.xlsx">下载选型结果</a>'
        st.markdown(html, unsafe_allow_html=True)

if __name__ == "__main__":
    main()


